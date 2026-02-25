"""
Build a summary Excel report from existing benchmark result CSVs.

Usage:
    python3 build_summary.py [--results-dir ./results] [--output ./results/summary.xlsx]
"""

import argparse
import glob
import os
import re
import subprocess
import tempfile

import pandas as pd
from tqdm import tqdm


SCORE_PREFIX = "score_"
LLM_OUTPUT_COL = "llm_response"
MMDC = os.environ.get("MMDC_PATH", "mmdc")


def strip_code_fences(text: str) -> str:
    """Strip markdown code fences, preserving sequenceDiagram header.

    Formats seen in the wild:
      ```mermaid\nsequenceDiagram\n...```   → strip ```mermaid, keep sequenceDiagram
      ```sequenceDiagram\n...```            → strip only ```, keep sequenceDiagram
      ```\nsequenceDiagram\n...```          → strip ```
    """
    if not text:
        return ""
    text = text.strip()
    text = re.sub(r"^```mermaid\s*\n?", "", text)
    text = re.sub(r"^```(?=sequenceDiagram)", "", text)
    text = re.sub(r"^```\s*\n", "", text)
    text = re.sub(r"\n?```\s*$", "", text)
    return text.strip()


def run_mmdc(code: str) -> bool:
    with tempfile.NamedTemporaryFile(mode="w", suffix=".mmd", delete=False) as f:
        f.write(code)
        mmd_path = f.name
    svg_path = mmd_path.replace(".mmd", ".svg")
    try:
        result = subprocess.run(
            [MMDC, "-i", mmd_path, "-o", svg_path],
            capture_output=True, text=True, timeout=30,
        )
        return result.returncode == 0
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return False
    finally:
        for p in (mmd_path, svg_path):
            if os.path.exists(p):
                os.unlink(p)


def process_csv(csv_path: str) -> dict | None:
    df = pd.read_csv(csv_path)
    score_cols = [c for c in df.columns if c.startswith(SCORE_PREFIX)]
    if not score_cols:
        return None

    model_name = os.path.basename(os.path.dirname(csv_path))
    total = len(df)
    row = {"model": model_name, "samples": total}

    for col in score_cols:
        criterion = col.removeprefix(SCORE_PREFIX)
        row[criterion] = round(df[col].mean(skipna=True), 3)

    row["judge_avg"] = round(df[score_cols].mean(axis=1).mean(skipna=True), 3)

    # Render rate via mmdc
    rendered = 0
    failed_rows = []
    for idx, r in tqdm(df.iterrows(), total=total, desc=f"[{model_name}] mmdc render"):
        code = strip_code_fences(str(r.get(LLM_OUTPUT_COL, "")))
        if code and run_mmdc(code):
            rendered += 1
        else:
            failed_rows.append(idx)
    if failed_rows:
        print(f"  [{model_name}] failed rows: {failed_rows}")
    row["render_rate"] = round(rendered / total, 3) if total > 0 else 0

    return row


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--results-dir", default="./results")
    parser.add_argument("--output", default="./results/summary.xlsx")
    args = parser.parse_args()

    csv_files = sorted(glob.glob(os.path.join(args.results_dir, "*/results__*.csv")))
    if not csv_files:
        print(f"No CSVs found in {args.results_dir}/*/")
        return

    rows = [process_csv(p) for p in csv_files]
    summary = pd.DataFrame([r for r in rows if r])
    summary = summary.sort_values("judge_avg", ascending=False)

    print("\n" + "=" * 90)
    print("MermaidSeqBench Summary")
    print("=" * 90)
    print(summary.to_string(index=False))
    print("=" * 90)

    # Описания колонок
    descriptions = {
        "model": "Название модели (OpenRouter ID без провайдера)",
        "samples": "Количество примеров из датасета MermaidSeqBench",
        "syntax": "Синтаксическая корректность MermaidJS: правильность participant, activate/deactivate, alt/else/end",
        "mermaid_only": "Содержит ли ответ только MermaidJS-код без лишнего текста и пояснений",
        "logic": "Логика и полнота потока: каждый запрос имеет ответ, альтернативные ветки покрыты",
        "completeness": "Полнота покрытия: все участники, пары запрос/ответ, точки решений из задания",
        "activation_handling": "Корректность activate/deactivate: баланс, отсутствие лишних deactivate",
        "error_and_status_tracking": "Обработка ошибок и отслеживание статусов: разделение success/failure потоков",
        "judge_avg": "Среднее по всем 6 критериям LLM-as-a-Judge (GPT-5.1)",
        "render_rate": "Доля диаграмм, которые успешно рендерятся через mmdc (Mermaid CLI)",
    }

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
        ws = writer.sheets["Summary"]

        # Заголовок
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary.columns))
        title_cell = ws.cell(row=1, column=1, value="MermaidSeqBench — Сравнение моделей")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Стиль заголовков колонок
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx in range(1, len(summary.columns) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Форматирование данных
        for col_idx, col_name in enumerate(summary.columns, start=1):
            for row_idx in range(3, 3 + len(summary)):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center")
                if col_name == "render_rate":
                    cell.number_format = "0.0%"
                elif col_name not in ("model", "samples"):
                    cell.number_format = "0.000"

        # Ширина колонок
        for col_idx, col_name in enumerate(summary.columns, start=1):
            width = max(len(str(col_name)) + 4, 14)
            if col_name == "model":
                width = 26
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Легенда — пояснения к колонкам
        legend_start = 3 + len(summary) + 2
        ws.cell(row=legend_start, column=1, value="Пояснения к колонкам:").font = Font(bold=True, size=11)
        legend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for i, col_name in enumerate(summary.columns):
            row = legend_start + 1 + i
            name_cell = ws.cell(row=row, column=1, value=col_name)
            name_cell.font = Font(bold=True)
            name_cell.fill = legend_fill
            desc_cell = ws.cell(row=row, column=2, value=descriptions.get(col_name, ""))
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(summary.columns))

    print(f"\nSaved to {args.output}")


if __name__ == "__main__":
    main()
